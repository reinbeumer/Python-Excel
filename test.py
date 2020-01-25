# -*- coding: utf-8 -*-
#!/usr/bin/env python3

#
# het runnen van dit script begint onderaan.
# daarvoor staan alle globale variabelen en functies die daarvoor nodig zijn.
#

# dit is de lib voor het lezen en schrijven van Excel spreadsheets
import openpyxl
# het relative pad naar het import bestand, moet naast het script staan (in de zelfde map)
spreadSheetname = "voorbeeld data.xlsx"
# open het werkboek (de Excel) met openpyxl en bewaar dit in de variabele 'workbook'
workbook = openpyxl.load_workbook(spreadSheetname)
# van de variabele 'workbook' bewaar het active blad in de variabelen 'sheet'
sheet = workbook.active

# hieronder een aantal globale variabelen die voor meerdere onderdeling in het script nodig zijn.
# de onbekende worden met aan standaard waarde van 0 gevuld

# maak een (dict) variabele met de vereiste kopnamen van het Excel bestand als sleutels (keys) en de kolom nummers als waardes (values)
# de waardes worden in 'setHeaders()' bepaald en gezet
headers = {"item":0, "formaat": 0, "aantal": 0, "prijs": 0}
# maak een (int) variabele waar de kop rij in kan worden opgeslagen
# de waarde wordt in 'getHeaderRow()' gezet
headerrow = 0
# maak een (dict) variable met de formaten als sleutels (keys) en de stuksprijs als waarde (value)
calc = {"A4": 0.25, "A3": 0.50, "A2": 1.00}
# maak een (float) variabele waar de totale waarde in kan worden opgeslagen
total = 0.0

# dit is de functie die de koppen rij opzoekt en eenmaal gevonden die in de variabele 'headerrow' zet
def getHeaderRow():
    # dit is de globale var die moet worden voorgegaan met 'global' anders kun je hem niet veranderen buiten deze functie
    global headerrow
    # dit is de lus (loop) die door alle rijen heen loopt
    for row in sheet.iter_cols():
        # dit is een teller die bijhoud hoeveel van de benodigde koppen gevonden zijn in een rij
        # bij iedere nieuwe rij moet deze weer op 0 gezet worden
        headerHitter = 0
        # dit is de binnen lus (inner loop) die door iedere cel loopt van de rij
        for cell in row:
            # hier controleer je of de waarde van de cel van de rij voor komt in de 'headers' dict
            if cell.value in headers:
                # als dat zo is dan moet de 'headerHitter' met 1 opgehoogd worden
                # de += betekend de huidige waarde + 1
                # dit is het zelfde als "headerHitter = headerHitter + 1", dit is korter
                headerHitter += 1
                # als de 'headerHitter' gelijk is aan het aantal benodigde koppen uit 'headers', dan is dat de koppen rij
                if headerHitter == headers.__len__():
                    # zet de rij waar de buitenste lus in staat als de kopenrij
                    headerrow = cell.row
                    # de kop is gevonden dus voeven we niet verder te zoeken
                    break

# dit is de functie die de kolom nummers opzoekt van de koppen uit 'headers' in de 'headerrow' die net gevonden is
# dit had ook in 1 keer worden gedaan met de 'getHeaderRow()' functie maar twee aparte stappen is makkelijker te volgen
def setHeaders():
    # dit is de globale var waar de kolomnummers van gevonden koppen worden opgeslagen
    # de volgorde maakt op deze manier niets maar uit, die zoeken we dus op
    global headers
    # de lus met een lengte van 1 (de 'headerrow') om de rij als rij object te pakken
    for row in sheet.iter_rows(headerrow):
        # de binnen lus om door de cellen te wandelen van de 'headerrow'
        for cell in row:
            # als de waarde van de cel voorkomt in de 'headers' dict dan is dat de kolom nummer
            if cell.value in headers:
                # zet het gevonden kolom nummer als waarde bij de sleutel van 'headers'
                # dan kun je deze altijd uit 'headers' opvragen met de naam
                headers[cell.value] = cell.column

# dit is de functie waar de berekeningen plaats vinden
def calcPrices():
    # dit is de globale var voor het vasthouden van het totaal bedrag
    global total
    # dit is de lus die door alle rijen heen loopt vanaf de eerst volgende rij na de koppen rij, vandaar 'headerrow+1'
    for row in sheet.iter_rows(min_row=headerrow + 1):
        # dir zijn de variabelen waar per rij de waardes van de benodigde cellen wordt opgeslagen
        # omdat we verschillende benodigde kolomen hebben opgeslagen in de 'headers' dict, moeten we die ook gebruiken om de kolom nummer op te halen van de gewenste kop
        # dus de 'formaatCel' waarde wordt uit de huidige rij gehaald met rijnummer 'row[0].row' en daarvan de kolom met het kolomnummer van "formaat" uit de 'headers' dict
        formaatCel = sheet.cell(column=headers.get("formaat"), row=row[0].row)
        # dit geldt ook voor het aantal
        aantalCel = sheet.cell(column=headers.get("aantal"), row=row[0].row)
        # en ook voor de prijs kolom, deze is eeg maar daar gaan wij de berekende waarde in stoppen
        prijsCel = sheet.cell(column=headers.get("prijs"), row=row[0].row)
        # als het aantal leeg wordt gelaten dan is de waarde 'None', dan zetten we hem op nul (0)
        if aantalCel.value == None:
            # op 0 zetten anders kunnen we niet rekenen
            aantalCel.value = 0
        # als het formaat voorkomt in de 'calc' dict, dan kunnen we wat uitrekenen
        if formaatCel.value in calc:
            # hier maken we een nieuwe var waar we de waarde uitrekenen door het aantal te vermenigvuldigen met de waarde die we hebben kunnen vinden in de 'calc' dict
            itemPrijs = aantalCel.value * calc.get(formaatCel.value)
            # met 'prijsCel.coordinate' kunnen we de coördinaten uit de sheet halen van de 'prijsCel'
            # en daar kunnen we de nieuwe waarde zetten die we zojuist hebben uitgerekend
            sheet[prijsCel.coordinate].value = itemPrijs
            # dit is een formating voor de eure, die heb ik gekopieerd uit de opmaak van Excel (Libre office in mijn geval)
            sheet[prijsCel.coordinate].number_format = "[$€-413] #.00;[$€-413] #.00-"
            # het totaal ophogen met de uitgerekende waarde
            total += itemPrijs
            # even een printje naar de console van de uitgerekende waarde
            # heeft verder geen waarde voor dit script, is meer voor jezelf om te zien of het een beetje klopt
            print(sheet[prijsCel.coordinate].value)

# dit is de functie waar we het totaal toevoegen onder aan de sheet
# de tekst 'totaal' komt onder de kolom 'item' en de waarde van het totaal onder de kolom 'prijs'
# en we gieten er nog wat opmaak overheen
def modFile():
    # hier kijken we wat de laatste rij is
    totalRow = sheet.max_row
    # hier maken we een var waar de coördinaten van de laatste rij van de item kolom komt te staan
    totalnameCoord = sheet.cell(column=headers.get("item"), row=totalRow).coordinate
    # als in deze laatste rij niet het woordje 'totaal' staat gaan we er 1 lager anders overschrijven we de laatste rij
    # als hier WEL het woordje 'totaal' staat, dan gaan we deze overschrijven
    # als we dit niet doen dan wordt bij iedere run een extra regel toegevoegd waar het totaal komt te staan
    if sheet[totalnameCoord].value != "totaal":
        # 1 lager want er staat geen 'totaal'
        totalRow += 1
        # de nieuwe coördinaten ophalen
        totalnameCoord = sheet.cell(column=headers.get("item"), row=totalRow).coordinate
    # Zet de waarde in de sheet met de coördinaten van 'totalnameCoord' op de tekst 'totaal'
    sheet[totalnameCoord].value = "totaal"
    # maak er nog wat moois van met wat topmaak
    sheet[totalnameCoord].font = openpyxl.styles.Font(bold=True, italic=True, name='Calibri')
    # maak een nieuwe var waar de coördinaten van het totaal wordt opgeslagen
    totalValCoord = sheet.cell(column=headers.get("prijs"), row=totalRow).coordinate
    # Zet de waarde in de sheet met de coördinaten van 'totalValCoord' op het uitgerekende bedrag die in de var 'total' staat
    sheet[totalValCoord].value = total
    # maak er nog wat moois van met wat topmaak
    sheet[totalValCoord].font = openpyxl.styles.Font(bold=True, italic=True, name='Calibri')
    # formateer het als een valuta
    sheet[totalValCoord].number_format = "[$€-413] #.00;[$€-413] #.00-"
    # zet er een dubbele lijn boven
    sheet[totalValCoord].border = openpyxl.styles.Border(top=(openpyxl.styles.Side(border_style='double', color='8BB048')))
    # schrijf alles terug in een nieuw bestand
    workbook.save(filename='MOD_'+spreadSheetname)

# hier begint het script met uitvoren van d everschillende stappen
# daarvoor waren het allen nog variabelen die in het geheugen zijn gezet
# een functie doet niks als je hem niet aanroept

# 1. dit is de eerste stap, de herder regel opzoeken
getHeaderRow()
# een printje naar de console voor het debuggen
print(headerrow)
# 2. dit is de tweede stap, het zetten van de kolom nummers bij de koppen met de gegevens van de vorige stap
setHeaders()
# ook weer een debug printje naar de console
print(headers)
# 3. dit is stap drie, het uitrekenen van de prijzen per regel en het totaal met de gegevens van de vorige stapen
calcPrices()
# ook weer een debug printje naar de console
print(total)
# 4. dit is de laatste stap, de totalen toevoegen en weg schrijven
modFile()
# een laatste printje zodat ik kan zien dat hij klaar is
# alle andere printjes kunnen weg maar deze laat ik altijd staan, dan weet ik wanner hij klaar is
print('DONE!!')
